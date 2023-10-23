import argparse
import logging
import textwrap
import pandas as pd
import openpyxl
import json

from failures.articles.models import Article, SearchQuery
from failures.networks.models import ChatGPT

from typing import List
from langchain.llms import OpenAI
from langchain.output_parsers import PydanticOutputParser
from langchain.prompts import PromptTemplate
from langchain.pydantic_v1 import BaseModel, Field, validator

class EvaluatePostmortemCommand:
    def prepare_parser(self, parser: argparse.ArgumentParser):
        """
        Prepare the argument parser for the evaluate postmortem command.

        Args:
            parser (argparse.ArgumentParser): The argument parser to configure.
        """
        # add description
        parser.description = textwrap.dedent(
            """
            Evaluate the performance of the LLM to construct a postmortem. This evaluation uses ChatGPT to 
            evaluate model outputs with reference to the manual answers. 
            If no arguments are provided, all articles that have been classified and postmortems created will be used
            to score performance. 
            If --list is provided then a list of all articles that did not match will be outputted


            Evaluate the performance of the LLM predicting whether or not a given article is a software failure.
            If no arguments are provided, all articles that have been classified will be used to score performance.
            If --all is provided then more metrics will be outputted (# Right, # Wrong, # False Positive, # False Negative,
            # Evaluated).
            If --articles is provided then a only the list of articles provided will be evaluated
            """
        )
        parser.add_argument(
            "--all",
            action="store_true",
            help="Lists all metrics.",
        )
        parser.add_argument(
            "--list",
            action="store_true",
            help="List all articles incorrectly classified.",
        )
        parser.add_argument(
            "--articles",
            nargs="+",  # Accepts one or more values
            type=int,    # Converts the values to integers
            help="A list of integers.",
        )

    def run(self, args: argparse.Namespace, parser: argparse.ArgumentParser):
        """
        Run the evaluation process based on the provided arguments.

        Args:
            args (argparse.Namespace): The parsed command-line arguments.
            parser (argparse.ArgumentParser): The argument parser used for configuration.

        """
        # Creating metrics to return
        metrics = {}

        # Define the path to your Excel file
        excel_file_path = "./tests/manual_evaluation/Pilot_Taxonomy_Manual.xlsx"

        # Load the Excel workbook
        workbook = openpyxl.load_workbook(excel_file_path)

        # Select the worksheet (assuming the sheet name is "Sheet1")
        worksheet = workbook['Sheet1']

        # Initialize a list to store dictionaries
        data_list = []

        # Iterate through columns starting from the 2nd column (1B, 1C, 1D, ...)
        col_num = 2
        while worksheet.cell(row=1, column=col_num).value is not None:
            data_dict = {}
            for row_num in range(2, 24):
                # Get the title from row 2 (2A, 3A, 4A, ...)
                title_cell = worksheet.cell(row=row_num, column=1)
                title = title_cell.value

                # Get the corresponding value from the current row
                value_cell = worksheet.cell(row=row_num, column=col_num)
                value = value_cell.value

                # Add the data to the dictionary
                data_dict[title] = value

            # Append the dictionary to the list
            data_list.append(data_dict)

            # Move to next column
            col_num += 1

        # Close the workbook
        workbook.close()

        if not data_list:
            logging.info("evaluate_postmortem: Could not fetch manual information.")
            return

        self.__compare_similarity_test(data_list[0], data_list[1])

    def __evaluate_response(self, response: dict):
        """
        Evalaute the ChatGPT similarity resposne

        Args:
            response (dict): LangChain formatted response for similarity metrics of all categories for an article
        """

    def __compare_similarity_test(self, manual: dict, automated: dict, specific: str = None):
        """
        Use ChatGPT to compare the similarity between the manual postmortem and
        the automated postmortem

        Args:
            manual (dict): The manual postmortem data
            automated (dict): The LLM generated postmortem data
            specifc (str, optional): Optional input parameter to add specific details to prompt
        """
        # Create a ChatGPT instance
        chatGPT = ChatGPT() 

        # Define the query
        original_query = """
        You will be provided with an answer denoted by \"Answer:\". Check if the following pieces of information denoted by \"Statements: \" are directly contained in the answer.
        For each of these points perform the following steps:
        """

        # Set up parser + inject instructions into prompt template
        parser = PydanticOutputParser(pydantic_object=SimilarQuery)

        prompt = PromptTemplate(
            template="Answer the user query.\n{format_instructions}\n{statement}\n{answer}\n",
            input_variables=["statement", "answer"],
            partial_variables={"format_instructions": parser.get_format_instructions()},
        )

        # Define set of categories to score
        categories = ["system", "time", "se-causes", "nse-causes", "impact", "mitigation"]

        # Defining return object
        response = {}

        # Score all categories
        for (manual_key, manual_val), (auto_key, auto_val) in zip(manual.items(), automated.items()):
            if manual_key not in categories:
                continue

            _input = prompt.format_prompt(statement="Statement: " + str(manual_val), answer="Answer: " + str(auto_val))

            # Define input data
            input_data = {
                "messages": [
                    {"role": "user", "content": _input.to_string()}
                ],
                "model": "gpt-3.5-turbo",
                "temperature": 0
            }

            # Get and parse output
            output = chatGPT.predict(input_data)
            parser.parse(output)
            response[manual_key] = json.loads(output)

        return response

class SimilarQuery(BaseModel):
    statement_in_answer: list = Field(description="1. List the components of the statement not in the answer. Return the statements in a Python list format.")
    answer_in_statement: list = Field(description="2. List the components of the answer not in the statement. Return the statements in a Python list format.")
    similar_bool: str = Field(description="Do both the statement and the answer contain the same information? Write \"yes\" if the answer is yes, otherwise write \"no\".")
    count: int = Field(description="Finally, provide a count of how many \"yes\" answers there are.")
